import os
import time
import inspect
import pathlib
from datetime import datetime
# If the application is already live, use SimpleTestCase
# To test by running the webserver, use LiveServerTestCase
from django.test import LiveServerTestCase, SimpleTestCase
"""
Better use reverse so the url can be called by their namespaces
from django.urls import reverse
import pyotp
"""
from django.contrib.auth.hashers import make_password
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support.select import Select
from openpyxl import Workbook, load_workbook
from .control_data import TestController


class LoginTestCase(SimpleTestCase, TestController):
    LOGIN_URL = '<DOMAIN HERE>' # reverse('index')
    login_container = None
    test_folder = None
    phone = None
    email = None
    password = None
    wb = None
    ws = None
    _input = None
    _name = None

    # Create excel file
    wb = Workbook()
    ws = wb.active
    ws['A1'] = 'Cases'
    ws['B1'] = 'Sample Input'
    ws['C1'] = 'Expected Result'
    ws['D1'] = 'Status'
    ws['E1'] = 'Image Name'
    wb.save(os.path.join(
        os.path.abspath('.'), 'LoginTestCase.xlsx'))
    
    def setUp(self):
        # Load worksheet
        self.wb = load_workbook(os.path.join(
            os.path.abspath('.'), self.__class__.__name__+'.xlsx'))
        # Create control data
        self.create_dummy()
        # Create folder for the image to be saved
        self.test_folder = os.path.join(
                        os.path.abspath('.'), self.__class__.__name__)
        pathlib.Path(self.test_folder).mkdir(exist_ok=True)
        # Maximize window
        options = webdriver.ChromeOptions()
        options.add_argument('--no-sandbox')
        options.add_argument("--remote-debugging-port=9222")
        options.add_argument("--start-maximized")
        # Due to the dev environment is Windows Subsystem Ubuntu, the webdriver cannot find
        # the chromedriver and open chrome, so it has to be declared in the Chrome() 
        # self.selenium = webdriver.Chrome(chrome_options=options)
        self.selenium = webdriver.Chrome('./lib/chromedriver.exe', chrome_options=options)
        selenium = self.selenium
        # Runs the webserver in local, it takes the server url and namespace
        # selenium.get('%s%s' % (self.live_server_url, self.LOGIN_URL))

        # Access the remote or live server
        selenium.get('%s' % (self.LOGIN_URL))
        
        # It is much easier to find the elements if it has a name.
        # 
        # self.login_container = selenium.find_element_by_id(
        #                         'loginFormContainer')
        # self.email = self.login_container.find_element_by_name('email')
        # self.password = self.login_container.find_element_by_name('password')
        # self.submit = selenium.find_element_by_id('loginBtn')

        # If there are NO name or id for the input boxes or buttons, use xpath
        # Using WebDriverWait, this checks if the element is ready, good for slow loading
        # websites
        self.email = WebDriverWait(selenium, 5).until(
            EC.visibility_of_element_located((By.XPATH, "//input[@placeholder='Mobile Number']")))
        self.password = selenium.find_element_by_xpath("//input[@placeholder='Password']")
        self.submit = WebDriverWait(selenium, 5).until(
            EC.visibility_of_element_located((By.XPATH, "//button[@type='submit']")))
        super(LoginTestCase, self).setUp()

    def tearDown(self):
        self.selenium.quit()
        super(LoginTestCase, self).tearDown()

    def __save_worksheet(self, case):
        """
        Updates the excel sheets with the test case results
        Use this one for tracking and distribution to QA

        :Parameters:
            case: (str) test case 
            _input: (str) sample input
            _name: (str) function name
        """
        row = self.ws.max_row + 1
        self.ws['A'+str(row)] = case
        self.ws['B'+str(row)] = self._input
        self.ws['D'+str(row)] = 'pass'
        self.ws['E'+str(row)] = self._name
        self.wb.save(os.path.join(
            os.path.abspath('.'), self.__class__.__name__+'.xlsx'))

    def test_phone_empty(self):
        self.email.send_keys('')
        self.submit.send_keys(Keys.RETURN)
        err_list = WebDriverWait(self.selenium, 5).until(
            EC.visibility_of_all_elements_located((By.XPATH, "//div[contains(@class, 'ant-form-explain')]")))
        expected_error = ['Phone number is required.', 'Password is required.']
        # Its unsure if the err_list returns the error in order of the input boxes
        # this may turnout false flag if it the error changes order but for sample sake
        # checking will be done like this.
        for err, exp_err in zip(err_list, expected_error):
            self.assertEqual(exp_err, err.text)
        self.save_image(inspect.stack()[0][3], self.test_folder, self.selenium)

    def test_login_valid(self):
        # self.user_email and password are located in control_data.py
        self.email.send_keys(self.user_email)
        self.password.send_keys(self.user_password)
        # Verify correct input
        self.save_image(inspect.stack()[0][3], self.test_folder, self.selenium)
        self.submit.send_keys(Keys.RETURN)
        # Check if user should be redirected to the next page after login
        # The ENDPOINT value is namespace in the urls.py for this implementation
        # ENDPOINT = '<DOMAIN HERE>'
        # wait = WebDriverWait(self.selenium, 5)
        # element = wait.until(EC.url_to_be(self.live_server_url + ENDPOINT))
        # OR
        # element = wait.until(EC.url_to_be(ENDPOINT))
        # self.assertTrue(element)
        # self.save_image(inspect.stack()[0][3], self.test_folder, self.selenium)

    # Continue adding the testcases until everything is covered.